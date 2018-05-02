# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import openpyxl

from openpyxl.utils import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
    range_boundaries,
    rows_from_range,
    coordinate_to_tuple,
    absolute_coordinate,
)

from portal.models import Record, Song, Artist, Company

import logging
import logging.config
from portal.excel.excel_logging import LOGGING

logger = logging.getLogger('xlsx')

ws_record_titles = ['唱片标题', '唱片监制', '唱片编号', '介质', '发布时间', '发布时间排序', '唱片公司', '唱片歌手', '唱片录音', '唱片混音', '唱片说明',
                    '唱片乐手信息', '序号', '歌手', '歌名', '作曲', '作词', '编曲', '歌曲乐手信息', '歌曲合唱', '歌曲和唱', '歌曲监制', '歌曲说明']
ws_artist_titles = ['歌手姓名', '歌手类型']


def str_strip(value):
    s = str(value) if value else None
    if s:
        s = s.strip()
    return s


class WSRowRecord:
    class WSRecord:
        title: str = None
        producer: str = None
        number: str = None
        format: str = None
        release: str = None
        release_order: str = None
        recorder: str = None
        mixer: str = None
        bandsman: str = None
        description: str = None

        company_name: str = None

        artists: str = None

        @staticmethod
        def read_from_row(row):
            record = WSRowRecord.WSRecord()

            record.title = str_strip(row[0].value)
            record.producer = row[1].value
            record.number = str_strip(row[2].value)
            record.format = row[3].value
            record.release = row[4].value
            record.release_order = row[5].value
            record.recorder = row[8].value
            record.mixer = row[9].value
            record.bandsman = row[11].value
            record.description = row[10].value

            record.company_name = row[6].value

            record.artists = row[7].value

            return record

    class WSSong:
        track: str = None
        title: str = None
        composer: str = None
        lyricist: str = None
        arranger: str = None
        bandsman: str = None
        vocalist: str = None
        producer: str = None
        description: str = None

        artists: str = None
        artists_part: str = None

        @staticmethod
        def read_from_row(row):
            song = WSRowRecord.WSSong()

            song.track = str_strip(row[12].value)
            song.title = str_strip(row[14].value)
            song.composer = row[15].value
            song.lyricist = row[16].value
            song.arranger = row[17].value
            song.bandsman = row[18].value
            song.vocalist = row[20].value
            song.producer = row[21].value
            song.description = row[22].value

            song.artists = row[13].value
            song.artists_part = row[19].value

            return song

    record: WSRecord = None
    song: WSSong = None

    @staticmethod
    def read_from_row(row):
        ws_row = WSRowRecord()

        ws_row.record = WSRowRecord.WSRecord.read_from_row(row)
        ws_row.song = WSRowRecord.WSSong.read_from_row(row)

        return ws_row


class WSRowArtist:
    name: str = None
    type: str = None

    @staticmethod
    def read_from_row(row):
        ws_row = WSRowArtist()

        ws_row.name = str_strip(row[0].value)
        ws_row.type = str_strip(row[1].value)

        return ws_row


class ExcelParser(object):
    def __init__(self, filename_excel, filename_log):

        self.filename_excel = filename_excel
        self.filename_log = filename_log

        LOGGING['handlers']['file']['filename'] = self.filename_log
        logging.config.dictConfig(LOGGING)

        logger.info("[Excel]初始化")

    def parse(self):
        logger.info("[Excel]开始加载")
        wb = openpyxl.load_workbook(filename=self.filename_excel, read_only=True)
        logger.info("[Excel]加载完毕")

        logger.info("[Excel]工作表数量: {0}".format(len(wb.worksheets)))
        logger.info("[Excel]工作表名称: {0}".format(wb.sheetnames))

        try:
            self.parse_worksheet_record(wb.worksheets[0])
            self.parse_worksheet_artist(wb.worksheets[1])
        except Exception as e:
            logger.error("[Excel]异常错误: {0}".format(e))

            import sys, traceback

            # traceback.print_exc()

            exc_type, exc_value, exc_traceback = sys.exc_info()
            str_list = traceback.format_exception(exc_type, exc_value, exc_traceback)

            msg = '\r\n'
            for str in str_list:
                msg += str
                msg += '\r\n'

            logger.error(msg)

        finally:
            logger.info("[Excel]导入结束")

    def parse_worksheet_record(self, ws):
        logger.info('[Record]处理开始')
        logger.info("[Record]工作表名称: {0}; 最大行数: {1}; 最大列数: {2};".format(ws.title, ws.max_row, ws.max_column))

        def _get_or_create_company(company_name):
            company_defaults = {
                'name': company_name
            }
            company = Company.objects.get_or_create(name=company_defaults['name'], defaults=company_defaults)[0]
            return company

        def _get_or_create_artist(artist_name):
            artist_defaults = {
                'name': artist_name,
            }
            artist = Artist.objects.get_or_create(name=artist_defaults['name'], defaults=artist_defaults)[0]
            return artist

        def _get_or_create_artists(artist_name_list):
            artists = []

            for artist_name in artist_name_list:
                artist = _get_or_create_artist(artist_name=artist_name)
                artists.append(artist)

            return artists

        def _save_record(row):
            ws_row = WSRowRecord.read_from_row(row)

            # Company
            company = None
            if ws_row.record.company_name:
                company = _get_or_create_company(company_name=ws_row.record.company_name)

            # Record Artists
            def xlxs_record_artist_name_list(cell):
                if cell:
                    return ws_row.record.artists.split('/')
                return None

            artist_name_list = xlxs_record_artist_name_list(ws_row.record.artists)

            artists = _get_or_create_artists(artist_name_list)

            record_defaults = {
                'title': ws_row.record.title,
                'producer': ws_row.record.producer,
                'number': ws_row.record.number,
                'format': Record.format_value_to_key(ws_row.record.format),
                'release': ws_row.record.release,
                'release_order': ws_row.record.release_order,
                'recorder': ws_row.record.recorder,
                'mixer': ws_row.record.mixer,
                'bandsman': ws_row.record.bandsman,
                'description': ws_row.record.description,
                'company': company,
            }

            obj, created = Record.objects.update_or_create(title=record_defaults['title'],
                                                           number=record_defaults['number'],
                                                           defaults=record_defaults)

            # print('record.artists', 'start', obj.artists, obj.artists.all())
            obj.artists.set(artists)
            # print('record.artists', 'stop', obj.artists, obj.artists.all())

            return obj

        def _save_record_song(record, row):
            ws_row = WSRowRecord.read_from_row(row)

            logger.info(
                '[{record_title}][{record_number}] [{song_track}][{song_title}]'.format(record_title=record.title,
                                                                                        record_number=record.number,
                                                                                        song_track=ws_row.song.track,
                                                                                        song_title=ws_row.song.title))

            # Song Artists

            def xlsx_song_artist_name_list(cell_artists, cell_artists_part):
                artist_name_list = []
                if cell_artists:
                    artist_name_list += cell_artists.split('/')
                if cell_artists_part:
                    artist_name_list += cell_artists_part.split('/')
                return artist_name_list

            artist_name_list = xlsx_song_artist_name_list(ws_row.song.artists, ws_row.song.artists_part)
            # print('song artist_name_list', artist_name_list)

            artists = _get_or_create_artists(artist_name_list)

            song_defaults = {
                'track': ws_row.song.track,
                'title': ws_row.song.title,
                'composer': ws_row.song.composer,
                'lyricist': ws_row.song.lyricist,
                'arranger': ws_row.song.arranger,
                'bandsman': ws_row.song.bandsman,
                'vocalist': ws_row.song.vocalist,
                'producer': ws_row.song.producer,
                'description': ws_row.song.description,
                'record': record
            }

            obj, created = Song.objects.update_or_create(track=song_defaults['track'], title=song_defaults['title'],
                                                         defaults=song_defaults)
            # print('song.artists', 'start', obj.artists, obj.artists.all())
            obj.artists.set(artists)
            # print('song.artists', 'stop', obj.artists, obj.artists.all())

            return obj

        ws_rows = iter(ws.rows)
        row0 = next(ws_rows)

        if not (row0[0].value == ws_record_titles[0] and row0[1].value == ws_record_titles[1] and row0[22].value ==
            ws_record_titles[22]):
            return

        # merged_cell_ranges_a = get_merged_cell_ranges_a(ws.merged_cell_ranges)
        # record_count = get_record_count(ws, merged_cell_ranges_a)
        # logger.info('专辑数量: {0}'.format(record_count))

        record = None
        record_size = 0

        for row in ws_rows:
            # col0 = row[0]
            # print(('col0', 'coordinate', 'row', 'column', 'col_idx'),
            #       (col0, col0.coordinate, col0.row, col0.column, col0.col_idx))

            ws_row = WSRowRecord.read_from_row(row)

            def should_new_record():
                if record is None:
                    return True
                else:
                    if ws_row.record.title and ws_row.record.title != record.title:
                        return True
                    elif ws_row.record.number and ws_row.record.number != record.number:
                        return True
                return False

            if should_new_record():
                record_size += 1
                logger.info('[{record_title}][{record_number}] ({count})'.format(count=record_size,
                                                                                 record_title=ws_row.record.title,
                                                                                 record_number=ws_row.record.number))
                record = _save_record(row)

            _save_record_song(record=record, row=row)

        logger.info('[Record]处理结束')

    def parse_worksheet_artist(self, ws):
        logger.info('[Artist]处理开始')
        logger.info("[Artist]工作表名称: {0}; 最大行数: {1}; 最大列数: {2};".format(ws.title, ws.max_row, ws.max_column))

        ws_rows = iter(ws.rows)
        row0 = next(ws_rows, None)

        if not (row0[0].value == ws_artist_titles[0] and row0[1].value == ws_artist_titles[1]):
            return

        artist_count = ws.max_row - 1
        logger.info('[Artist]歌手数量: {0}'.format(artist_count))

        for row in ws_rows:
            ws_row = WSRowArtist.read_from_row(row)

            # print(('col0', 'coordinate', 'row', 'column', 'col_idx'),
            #       (col0, col0.coordinate, col0.row, col0.column, col0.col_idx))

            artist_defaults = {
                'name': ws_row.name,
                'type': Artist.type_value_to_key(ws_row.type)
            }

            Artist.objects.update_or_create(name=artist_defaults['name'], defaults=artist_defaults)

        logger.info('[Artist]处理结束')
