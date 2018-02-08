import openpyxl

filename = 'document_template.xltx'
# filename = '/Users/liuxue/Downloads/Developer/GoldenTimes/唱片资料20170425.xlsx'
filename = '/Users/liuxue/Downloads/Developer/GoldenTimes/唱片资料20170425的副本.xlsx'

# wb = openpyxl.load_workbook(filename=filename, read_only=True)
wb = openpyxl.load_workbook(filename=filename)
print(wb.get_sheet_names())
print("The number of worksheets is {0}".format(len(wb.worksheets)))
print("Worksheet name(s): {0}".format(wb.get_sheet_names()))
ws = wb.worksheets[0]
print(ws)
print("{0} {1} {2}".format(ws.title, ws.max_row, ws.max_column))
print("Cell D30 is {0}".format(ws.cell(row=30, column=4).value))
# from openpyxl.utils.cell import range_boundaries, range_to_tuple, get_column_letter, rows_from_range
# [['唱片标题', '唱片监制', '唱片编号', '介质', '发布时间', '发布时间排序', '唱片公司', '唱片歌手', '唱片录音', '唱片混音', '唱片说明', '唱片乐手信息', '序号', '歌手', '歌名', '作曲', '作词', '编曲', '歌曲乐手信息', '歌曲合唱', '歌曲和唱', '歌曲监制', '歌曲说明']]

print(ws.rows)
for row in ws.rows:
    content = []

    line = [col.value for col in row]
    content.append(line)

    print(content)

# 太耗时
# print('merged_cells', ws.merged_cells)
print('merged_cell_ranges', ws.merged_cell_ranges)

from openpyxl.worksheet.merge import MergeCell, MergeCells
from openpyxl.cell.cell import Cell

# from openpyxl.worksheet.

# (3 -> 'C')
get_column_letter = openpyxl.cell.cell.get_column_letter(3)
print(get_column_letter)

# ('A' -> 1)
column_index_from_string = openpyxl.cell.cell.column_index_from_string('A')
print(column_index_from_string)

# openpyxl.cell.cell.xxx

print('=====')

ws = wb.worksheets[0]

print(ws)
print("{0} {1} {2}".format(ws.title, ws.max_row, ws.max_column))
print("Cell D30 is {0}".format(ws.cell(row=30, column=4).value))

print(ws.cell(row=1, column=1))
print('=====')

range_string = 'A2:A17'
min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(range_string)
print((min_col, min_row, max_col, max_row))

print(openpyxl.utils.cell.absolute_coordinate(range_string))
print(openpyxl.utils.cell.rows_from_range(range_string))

for row in openpyxl.utils.cell.rows_from_range(range_string):
    print(row)

tuple = openpyxl.utils.cell.range_to_tuple(range_string)
print(tuple)



# from portal.models import Song




